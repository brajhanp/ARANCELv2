from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.models import User 
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from .forms import HistorialBusquedaFilterForm, ReporteFilterForm, EditProfileForm, ChangePasswordForm
from django.contrib.auth.forms import UserCreationForm
from arancel.models import Seccion, Capitulo, Partida, Subpartida
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator
from .models import HistorialBusqueda, Rol, PerfilUsuario, Reporte, Bitacora
from django.db.models import Avg, Max, Min
from .forms import RegistroUsuarioForm
from django.contrib.auth.decorators import permission_required
from django.contrib.auth import update_session_auth_hash

def es_superusuario_o_tiene_permiso_usuarios(user):
    return user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol and user.perfil.rol.permisos_usuarios)

def es_administrador(user):
    """Verifica si el usuario es administrador (rol Administrador o superusuario)"""
    if user.is_superuser:
        return True
    if hasattr(user, 'perfil') and user.perfil.rol:
        return user.perfil.rol.nombre == 'Administrador' or user.perfil.rol.permisos_admin
    return False

def registrar_bitacora(request, tipo_accion, descripcion, detalles_adicionales=None):
    """Función helper para registrar acciones en la bitácora"""
    try:
        usuario = request.user if request.user.is_authenticated else None
        ip_address = None
        user_agent = None
        
        if request:
            # Obtener IP del cliente
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')
            
            # Obtener User-Agent
            user_agent = request.META.get('HTTP_USER_AGENT', '')[:255]
        
        Bitacora.objects.create(
            usuario=usuario,
            tipo_accion=tipo_accion,
            descripcion=descripcion,
            ip_address=ip_address,
            user_agent=user_agent,
            detalles_adicionales=detalles_adicionales or {}
        )
    except Exception as e:
        # No fallar si hay error al registrar bitácora
        import logging
        logging.error(f"Error al registrar en bitácora: {str(e)}")

@login_required
def home(request):
    total_secciones = Seccion.objects.count()
    total_capitulos = Capitulo.objects.count()
    total_partidas = Partida.objects.count()
    total_subpartidas = Subpartida.objects.count()
    
    # Obtener historial de búsquedas del usuario (últimas 10)
    historial = HistorialBusqueda.objects.filter(usuario=request.user)[:10]
    
    return render(request, 'home.html', {
        'total_secciones': total_secciones,
        'total_capitulos': total_capitulos,
        'total_partidas': total_partidas,
        'total_subpartidas': total_subpartidas,
        'historial_busquedas': historial,
    })

@login_required
def limpiar_historial(request):
    if request.method == 'POST':
        HistorialBusqueda.objects.filter(usuario=request.user).delete()
        messages.success(request, 'Historial de búsquedas limpiado correctamente.')
    return redirect('central:home')

@login_required
@user_passes_test(es_superusuario_o_tiene_permiso_usuarios)
def gestionar_usuarios(request):
    # Registrar acceso a gestión de usuarios en bitácora
    registrar_bitacora(
        request,
        'consulta',
        f'Acceso a gestión de usuarios por administrador {request.user.username}',
        {'accion': 'gestionar_usuarios', 'usuario_admin': request.user.username}
    )
    usuarios = User.objects.all().order_by('username')
    roles = Rol.objects.all().order_by('nombre')
    
    if request.method == 'POST':
        usuario_id = request.POST.get('usuario_id')
        rol_id = request.POST.get('rol_id')
        accion = request.POST.get('accion')
        
        if accion == 'asignar_rol':
            usuario = get_object_or_404(User, id=usuario_id)
            rol = get_object_or_404(Rol, id=rol_id) if rol_id else None
            
            perfil, created = PerfilUsuario.objects.get_or_create(usuario=usuario)
            perfil.rol = rol
            perfil.save()
            
            # Registrar en bitácora
            registrar_bitacora(
                request,
                'modificacion',
                f'Administrador {request.user.username} asignó rol "{rol.nombre if rol else "Sin rol"}" a usuario {usuario.username}',
                {'accion': 'asignar_rol', 'usuario_afectado': usuario.username, 'rol_asignado': rol.nombre if rol else None}
            )
            
            messages.success(request, f'Rol "{rol.nombre if rol else "Sin rol"}" asignado a {usuario.username}')
            
        elif accion == 'activar_desactivar':
            usuario = get_object_or_404(User, id=usuario_id)
            perfil, created = PerfilUsuario.objects.get_or_create(usuario=usuario)
            estado_anterior = perfil.activo
            perfil.activo = not perfil.activo
            perfil.save()
            
            estado = "activado" if perfil.activo else "desactivado"
            
            # Registrar en bitácora
            registrar_bitacora(
                request,
                'modificacion',
                f'Administrador {request.user.username} {estado} usuario {usuario.username}',
                {'accion': 'activar_desactivar', 'usuario_afectado': usuario.username, 'estado_anterior': estado_anterior, 'estado_nuevo': perfil.activo}
            )
            
            messages.success(request, f'Usuario {usuario.username} {estado}')
        elif accion == 'eliminar_usuario':
            usuario = get_object_or_404(User, id=usuario_id)
            if usuario.is_superuser:
                messages.error(request, 'No puedes eliminar un superusuario desde esta interfaz.')
            else:
                usuario_nombre = usuario.username
                usuario.delete()
                
                # Registrar en bitácora
                registrar_bitacora(
                    request,
                    'eliminacion',
                    f'Administrador {request.user.username} eliminó usuario {usuario_nombre}',
                    {'accion': 'eliminar_usuario', 'usuario_eliminado': usuario_nombre}
                )
                
                messages.success(request, f'Usuario {usuario_nombre} eliminado correctamente.')
    
    return render(request, 'central/gestionar_usuarios.html', {
        'usuarios': usuarios,
        'roles': roles,
    })

@login_required
@user_passes_test(es_superusuario_o_tiene_permiso_usuarios)
def gestionar_roles(request):
    roles = Rol.objects.all().order_by('nombre')
    
    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        if accion == 'crear_rol':
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion', '')
            permisos_arancel = request.POST.get('permisos_arancel') == 'on'
            permisos_admin = request.POST.get('permisos_admin') == 'on'
            permisos_usuarios = request.POST.get('permisos_usuarios') == 'on'
            
            if nombre:
                # Verificar que no se duplique un rol del sistema
                if nombre.lower() in ['administrador', 'despachador de aduana']:
                    messages.error(request, f'El rol "{nombre}" ya existe en el sistema.')
                else:
                    rol = Rol.objects.create(
                        nombre=nombre,
                        descripcion=descripcion,
                        permisos_arancel=permisos_arancel,
                        permisos_admin=permisos_admin,
                        permisos_usuarios=permisos_usuarios
                    )
                    
                    # Registrar en bitácora
                    registrar_bitacora(
                        request,
                        'creacion',
                        f'Administrador {request.user.username} creó rol "{rol.nombre}"',
                        {'accion': 'crear_rol', 'rol_creado': rol.nombre, 'permisos': {'arancel': permisos_arancel, 'admin': permisos_admin, 'usuarios': permisos_usuarios}}
                    )
                    
                    messages.success(request, f'Rol "{rol.nombre}" creado exitosamente')
            else:
                messages.error(request, 'El nombre del rol es obligatorio')
                
        elif accion == 'editar_rol':
            rol_id = request.POST.get('rol_id')
            rol = get_object_or_404(Rol, id=rol_id)
            
            # Permitir editar descripción y permisos de roles del sistema, pero no el nombre
            if rol.nombre in ['Administrador', 'Despachador de Aduana']:
                rol.descripcion = request.POST.get('descripcion', rol.descripcion)
                rol.permisos_arancel = request.POST.get('permisos_arancel') == 'on'
                rol.permisos_admin = request.POST.get('permisos_admin') == 'on'
                rol.permisos_usuarios = request.POST.get('permisos_usuarios') == 'on'
                rol.save()
                
                # Registrar en bitácora
                registrar_bitacora(
                    request,
                    'modificacion',
                    f'Administrador {request.user.username} actualizó permisos del rol "{rol.nombre}"',
                    {'accion': 'editar_rol', 'rol': rol.nombre, 'permisos': {'arancel': rol.permisos_arancel, 'admin': rol.permisos_admin, 'usuarios': rol.permisos_usuarios}}
                )
                
                messages.success(request, f'Permisos del rol "{rol.nombre}" actualizados exitosamente')
            else:
                # Para roles personalizados, permitir editar todo
                rol.nombre = request.POST.get('nombre', rol.nombre)
                rol.descripcion = request.POST.get('descripcion', rol.descripcion)
                rol.permisos_arancel = request.POST.get('permisos_arancel') == 'on'
                rol.permisos_admin = request.POST.get('permisos_admin') == 'on'
                rol.permisos_usuarios = request.POST.get('permisos_usuarios') == 'on'
                rol.save()
                
                # Registrar en bitácora
                registrar_bitacora(
                    request,
                    'modificacion',
                    f'Administrador {request.user.username} actualizó rol "{rol.nombre}"',
                    {'accion': 'editar_rol', 'rol': rol.nombre, 'permisos': {'arancel': rol.permisos_arancel, 'admin': rol.permisos_admin, 'usuarios': rol.permisos_usuarios}}
                )
                
                messages.success(request, f'Rol "{rol.nombre}" actualizado exitosamente')
            
        elif accion == 'eliminar_rol':
            rol_id = request.POST.get('rol_id')
            rol = get_object_or_404(Rol, id=rol_id)
            
            # Proteger roles del sistema
            if rol.nombre in ['Administrador', 'Despachador de Aduana']:
                messages.error(request, f'No se puede eliminar el rol "{rol.nombre}" porque es un rol del sistema.')
            else:
                # Verificar si hay usuarios usando este rol
                usuarios_con_rol = PerfilUsuario.objects.filter(rol=rol).count()
                if usuarios_con_rol > 0:
                    messages.error(request, f'No se puede eliminar el rol "{rol.nombre}" porque {usuarios_con_rol} usuario(s) lo están usando')
                else:
                    rol_nombre = rol.nombre
                    rol.delete()
                    
                    # Registrar en bitácora
                    registrar_bitacora(
                        request,
                        'eliminacion',
                        f'Administrador {request.user.username} eliminó rol "{rol_nombre}"',
                        {'accion': 'eliminar_rol', 'rol_eliminado': rol_nombre}
                    )
                    
                    messages.success(request, f'Rol "{rol_nombre}" eliminado exitosamente')
    
    return render(request, 'central/gestionar_roles.html', {
        'roles': roles,
    })

def register_view(request):
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Crear perfil de usuario por defecto
            PerfilUsuario.objects.create(usuario=user)
            messages.success(request, 'Usuario registrado correctamente. Ahora puedes iniciar sesión.')
            return redirect('login')
        else:
            messages.error(request, 'Error en el registro. Por favor, verifica los datos.')
    else:
        form = RegistroUsuarioForm()
    return render(request, 'registration/register.html', {'form': form})

def login_view(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Verificar si el usuario está activo
            try:
                perfil = user.perfil
                if not perfil.activo:
                    messages.error(request, 'Tu cuenta ha sido desactivada. Contacta al administrador.')
                    return render(request, 'registration/login.html')
            except PerfilUsuario.DoesNotExist:
                # Si no tiene perfil, crear uno por defecto
                PerfilUsuario.objects.create(usuario=user)
            
            login(request, user)
            
            # Registrar en bitácora
            registrar_bitacora(
                request,
                'login',
                f'Usuario {user.username} inició sesión exitosamente',
                {'username': user.username, 'user_id': user.id}
            )
            
            return redirect('home')
        else:
            # Registrar intento de login fallido
            registrar_bitacora(
                request,
                'login',
                f'Intento de login fallido para usuario: {username}',
                {'username': username, 'resultado': 'fallido'}
            )
            messages.error(request, 'Usuario o contraseña incorrectos')
    return render(request, 'registration/login.html')

def logout_view(request):
    # Registrar en bitácora antes de hacer logout
    if request.user.is_authenticated:
        registrar_bitacora(
            request,
            'logout',
            f'Usuario {request.user.username} cerró sesión',
            {'username': request.user.username, 'user_id': request.user.id}
        )
    logout(request) 
    return redirect('login')    

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario registrado correctamente. Ahora puedes iniciar sesión.')
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})    

@login_required
@user_passes_test(es_superusuario_o_tiene_permiso_usuarios)
def panel_admin_simplificado(request):
    """Panel de administración simplificado para usuarios con permisos de admin"""
    # Obtener estadísticas
    total_usuarios = User.objects.count()
    usuarios_activos = PerfilUsuario.objects.filter(activo=True).count()
    total_roles = Rol.objects.count()
    total_busquedas = HistorialBusqueda.objects.count()
    
    return render(request, 'central/panel_admin_simplificado.html', {
        'total_usuarios': total_usuarios,
        'usuarios_activos': usuarios_activos,
        'total_roles': total_roles,
        'total_busquedas': total_busquedas,
    })

@login_required
def admin_redirect(request):
    """Redirige a usuarios según sus permisos"""
    if request.user.is_superuser:
        # Superusuarios van al admin completo
        return redirect('/admin/')
    elif hasattr(request.user, 'perfil') and request.user.perfil.rol and request.user.perfil.rol.permisos_admin:
        # Usuarios con permisos admin van al panel simplificado
        return redirect('panel_admin_simplificado')
    else:
        # Usuarios sin permisos van al home
        messages.warning(request, 'No tienes permisos para acceder al panel de administración.')
        return redirect('home')    

class HistorialBusquedaListView(LoginRequiredMixin, ListView):
    model = HistorialBusqueda
    template_name = 'central/historial_busquedas.html'
    context_object_name = 'historial'
    paginate_by = 20
    
    def get_queryset(self):
        """
        Todos los usuarios (incluyendo administradores) ven solo su propio historial
        Para ver el historial de otros usuarios, usar la bitácora (solo para administradores)
        """
        queryset = HistorialBusqueda.objects.all()

        # Determinar si el solicitante es gerente o superusuario
        es_gerente = False
        if self.request.user.is_superuser:
            es_gerente = True
        elif hasattr(self.request.user, 'perfil') and self.request.user.perfil.rol:
            es_gerente = self.request.user.perfil.rol.nombre == 'Gerente'

        # Si no es gerente ni superusuario, ver solo su propio historial
        if not es_gerente:
            queryset = queryset.filter(usuario=self.request.user)
        else:
            # Gerentes y superusuarios pueden filtrar por usuario (campo opcional en el formulario)
            usuario_id = self.request.GET.get('usuario', '')
            if usuario_id:
                queryset = queryset.filter(usuario__id=usuario_id)

        # Aplicar filtros del formulario
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')
        tipo_resultado = self.request.GET.get('tipo_resultado', '')
        palabra_clave = self.request.GET.get('palabra_clave', '')
        
        if fecha_inicio:
            from django.utils.dateparse import parse_date
            fecha_parsed = parse_date(fecha_inicio)
            if fecha_parsed:
                queryset = queryset.filter(fecha_busqueda__date__gte=fecha_parsed)
        
        if fecha_fin:
            from django.utils.dateparse import parse_date
            fecha_parsed = parse_date(fecha_fin)
            if fecha_parsed:
                queryset = queryset.filter(fecha_busqueda__date__lte=fecha_parsed)
        
        if tipo_resultado:
            queryset = queryset.filter(tipo_resultado=tipo_resultado)
        
        if palabra_clave:
            queryset = queryset.filter(termino_busqueda__icontains=palabra_clave)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pasar el formulario de filtros al contexto
        context['form'] = HistorialBusquedaFilterForm(self.request.GET or None)
        return context

@login_required
@user_passes_test(es_administrador)
def bitacora_view(request):
    """Vista para ver la bitácora de auditoría (solo administradores)"""
    from .models import Bitacora
    from django.db.models import Q
    
    # Filtros
    tipo_accion = request.GET.get('tipo_accion', '')
    usuario_filtro = request.GET.get('usuario', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Obtener todas las entradas de bitácora
    bitacora = Bitacora.objects.all()
    
    # Aplicar filtros
    if tipo_accion:
        bitacora = bitacora.filter(tipo_accion=tipo_accion)
    
    if usuario_filtro:
        bitacora = bitacora.filter(usuario__username__icontains=usuario_filtro)
    
    if fecha_inicio:
        from django.utils.dateparse import parse_date
        fecha_inicio_parsed = parse_date(fecha_inicio)
        if fecha_inicio_parsed:
            from django.utils import timezone
            bitacora = bitacora.filter(fecha_accion__date__gte=fecha_inicio_parsed)
    
    if fecha_fin:
        from django.utils.dateparse import parse_date
        fecha_fin_parsed = parse_date(fecha_fin)
        if fecha_fin_parsed:
            bitacora = bitacora.filter(fecha_accion__date__lte=fecha_fin_parsed)
    
    # Ordenar por fecha más reciente
    bitacora = bitacora.order_by('-fecha_accion')
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(bitacora, 50)  # 50 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'central/bitacora.html', {
        'bitacora': page_obj,
        'tipo_accion': tipo_accion,
        'usuario_filtro': usuario_filtro,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'tipos_accion': Bitacora.TIPO_ACCION_CHOICES,
    })

@login_required
def exportar_historial(request):
    """Exporta historial de búsquedas a PDF o Excel"""
    format_type = request.GET.get('format', 'pdf')
    
    # Crear reporte de trazabilidad para la descarga del historial
    # Pero verificar si ya existe uno igual en el último segundo para evitar duplicados
    try:
        ahora = timezone.now()
        hace_un_segundo = ahora - timezone.timedelta(seconds=1)
        
        # Verificar si ya existe un reporte igual reciente
        reporte_reciente = Reporte.objects.filter(
            usuario=request.user,
            codigo_arancelario='HISTORIAL',
            tipo_accion='descarga_doc',
            fecha_operacion__gte=hace_un_segundo
        ).exists()
        
        # Solo crear si no existe uno reciente
        if not reporte_reciente:
            Reporte.objects.create(
                usuario=request.user,
                codigo_arancelario='HISTORIAL',
                descripcion_clasificacion=f'Descarga de historial de búsquedas en formato {format_type.upper()}',
                tipo_accion='descarga_doc',
                resultado_operacion='exitosa',
                detalles_adicionales=f'Usuario descargó historial de búsquedas en formato {format_type}'
            )
    except Exception:
        pass  # Si falla, continúa con la descarga normalmente
    
    # Obtener los datos filtrados
    queryset = HistorialBusqueda.objects.all()
    
    # Todos los usuarios (incluyendo administradores) solo ven su propio historial
    queryset = queryset.filter(usuario=request.user)
    
    form = HistorialBusquedaFilterForm(request.GET)
    
    if form.is_valid():
        if form.cleaned_data['fecha_inicio']:
            queryset = queryset.filter(fecha_busqueda__date__gte=form.cleaned_data['fecha_inicio'])
        if form.cleaned_data['fecha_fin']:
            queryset = queryset.filter(fecha_busqueda__date__lte=form.cleaned_data['fecha_fin'])
        if form.cleaned_data['tipo_resultado']:
            queryset = queryset.filter(tipo_resultado=form.cleaned_data['tipo_resultado'])
        if form.cleaned_data['palabra_clave']:
            queryset = queryset.filter(
                Q(termino_busqueda__icontains=form.cleaned_data['palabra_clave']) |
                Q(usuario__username__icontains=form.cleaned_data['palabra_clave'])
            )

    if format_type == 'excel':
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial de Búsquedas"

        # Agregar encabezados
        headers = ['Usuario', 'Término de búsqueda', 'Tipo de resultado', 'Fecha y hora']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Agregar datos
        for row, registro in enumerate(queryset, 2):
            ws.cell(row=row, column=1).value = registro.usuario.username
            ws.cell(row=row, column=2).value = registro.termino_busqueda
            ws.cell(row=row, column=3).value = registro.tipo_resultado or '--'
            ws.cell(row=row, column=4).value = registro.fecha_busqueda.strftime('%d/%m/%Y %H:%M')

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=historial_busquedas.xlsx'
        wb.save(response)
        return response

    else:  # PDF
        # Crear el documento PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Estilo para el título
        styles = getSampleStyleSheet()
        elements.append(Paragraph('Historial de Búsquedas', styles['Heading1']))
        elements.append(Paragraph(f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))

        # Datos para la tabla
        data = [['Usuario', 'Término de búsqueda', 'Tipo de resultado', 'Fecha y hora']]
        for registro in queryset:
            data.append([
                registro.usuario.username,
                registro.termino_busqueda,
                registro.tipo_resultado or '--',
                registro.fecha_busqueda.strftime('%d/%m/%Y %H:%M')
            ])

        # Crear la tabla
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

        # Generar el PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        # Crear la respuesta HTTP con el archivo PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=historial_busquedas.pdf'
        response.write(pdf)
        return response

@login_required
@login_required
def estadisticas_gravamenes(request):
    """Vista para mostrar estadísticas de gravámenes arancelarios por capítulo"""
    
    # Crear reporte de trazabilidad cuando acceden a gravámenes
    try:
        Reporte.objects.create(
            usuario=request.user,
            codigo_arancelario='',
            descripcion_clasificacion='Acceso a apartado de Estadísticas de Gravámenes',
            tipo_accion='consulta_detalle',
            resultado_operacion='exitosa',
            detalles_adicionales='Usuario consultó estadísticas de gravámenes arancelarios'
        )
    except Exception:
        pass  # No bloquea si falla
    
    # Obtener todos los capítulos
    capitulos = Capitulo.objects.all().order_by('codigo')
    
    # Lista para almacenar las estadísticas por capítulo
    estadisticas_por_capitulo = []
    
    for capitulo in capitulos:
        # Obtener todas las subpartidas de este capítulo a través de las partidas
        # Filtrar solo aquellas que tienen GA (no nulo)
        subpartidas = Subpartida.objects.filter(
            partida__capitulo=capitulo,
            ga__isnull=False
        )
        
        if subpartidas.exists():
            # Calcular estadísticas
            estadisticas = subpartidas.aggregate(
                promedio=Avg('ga'),
                maximo=Max('ga'),
                minimo=Min('ga')
            )
            
            # Agregar información del capítulo y estadísticas
            estadisticas_por_capitulo.append({
                'capitulo': capitulo,
                'promedio': estadisticas['promedio'],
                'maximo': estadisticas['maximo'],
                'minimo': estadisticas['minimo'],
                'cantidad_subpartidas': subpartidas.count()
            })
    
    return render(request, 'central/estadisticas_gravamenes.html', {
        'estadisticas': estadisticas_por_capitulo,
    })


# ============= VISTAS PARA REPORTES DE TRAZABILIDAD =============

class ReporteListView(LoginRequiredMixin, ListView):
    """Lista reportes de trazabilidad con filtros avanzados"""
    model = Reporte
    template_name = 'central/reporte_list.html'
    context_object_name = 'reportes'
    paginate_by = 50

    def get_queryset(self):
        queryset = Reporte.objects.all().select_related('usuario')
        
        # Verificar permisos: solo gerentes y superusuarios ven todos los reportes
        es_gerente = False
        if self.request.user.is_superuser:
            es_gerente = True
        elif hasattr(self.request.user, 'perfil') and self.request.user.perfil.rol:
            es_gerente = self.request.user.perfil.rol.nombre == 'Gerente'
        
        # Si no es gerente ni superusuario, solo ver sus propios reportes
        if not es_gerente:
            queryset = queryset.filter(usuario=self.request.user)
        
        form = ReporteFilterForm(self.request.GET)
        
        if form.is_valid():
            if form.cleaned_data.get('fecha_inicio'):
                queryset = queryset.filter(fecha_operacion__date__gte=form.cleaned_data['fecha_inicio'])
            if form.cleaned_data.get('fecha_fin'):
                queryset = queryset.filter(fecha_operacion__date__lte=form.cleaned_data['fecha_fin'])
            if form.cleaned_data.get('codigo_arancelario'):
                queryset = queryset.filter(codigo_arancelario__icontains=form.cleaned_data['codigo_arancelario'])
            if form.cleaned_data.get('tipo_accion'):
                queryset = queryset.filter(tipo_accion=form.cleaned_data['tipo_accion'])
            if form.cleaned_data.get('resultado_operacion'):
                queryset = queryset.filter(resultado_operacion=form.cleaned_data['resultado_operacion'])
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ReporteFilterForm(self.request.GET)
        context['total_reportes'] = self.get_queryset().count()
        return context


@login_required
def exportar_reportes(request):
    """Exporta reportes de trazabilidad a PDF o Excel.

    - Usuarios normales: solo pueden exportar sus propios reportes.
    - Gerentes y superusuarios: pueden exportar todos los reportes.
    """
    queryset = Reporte.objects.all().select_related('usuario')

    # Verificar permisos: solo gerentes y superusuarios ven todos los reportes
    es_gerente = False
    if request.user.is_superuser:
        es_gerente = True
    elif hasattr(request.user, 'perfil') and request.user.perfil.rol:
        es_gerente = request.user.perfil.rol.nombre == 'Gerente'

    # Si no es gerente ni superusuario, solo mostrar sus propios reportes
    if not es_gerente:
        queryset = queryset.filter(usuario=request.user)

    form = ReporteFilterForm(request.GET)
    if form.is_valid():
        if form.cleaned_data.get('fecha_inicio'):
            queryset = queryset.filter(fecha_operacion__date__gte=form.cleaned_data['fecha_inicio'])
        if form.cleaned_data.get('fecha_fin'):
            queryset = queryset.filter(fecha_operacion__date__lte=form.cleaned_data['fecha_fin'])
        if form.cleaned_data.get('codigo_arancelario'):
            queryset = queryset.filter(codigo_arancelario__icontains=form.cleaned_data['codigo_arancelario'])
        if form.cleaned_data.get('tipo_accion'):
            queryset = queryset.filter(tipo_accion=form.cleaned_data['tipo_accion'])
        if form.cleaned_data.get('resultado_operacion'):
            queryset = queryset.filter(resultado_operacion=form.cleaned_data['resultado_operacion'])

    format_type = request.GET.get('format', 'pdf')

    # Excel export
    if format_type == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reportes de Trazabilidad"

        headers = ['Fecha', 'Hora', 'Usuario', 'Código Arancelario', 'Descripción', 'Tipo de Acción', 'Resultado', 'Detalles']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        for row, reporte in enumerate(queryset, 2):
            ws.cell(row=row, column=1).value = reporte.fecha_operacion.strftime('%d/%m/%Y')
            ws.cell(row=row, column=2).value = reporte.fecha_operacion.strftime('%H:%M:%S')
            ws.cell(row=row, column=3).value = reporte.usuario.username
            ws.cell(row=row, column=4).value = reporte.codigo_arancelario or '--'
            ws.cell(row=row, column=5).value = reporte.descripcion_clasificacion or '--'
            ws.cell(row=row, column=6).value = dict(reporte.TIPO_ACCION_CHOICES).get(reporte.tipo_accion, reporte.tipo_accion)
            ws.cell(row=row, column=7).value = dict(reporte._meta.get_field('resultado_operacion').choices).get(reporte.resultado_operacion)
            ws.cell(row=row, column=8).value = reporte.detalles_adicionales or ''

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reportes_trazabilidad.xlsx'
        wb.save(response)
        return response

    # PDF export
    data = []
    headers = ['Fecha', 'Hora', 'Usuario', 'Código Arancelario', 'Descripción', 'Tipo de Acción', 'Resultado', 'Detalles']
    data.append(headers)
    for reporte in queryset:
        row = [
            reporte.fecha_operacion.strftime('%d/%m/%Y'),
            reporte.fecha_operacion.strftime('%H:%M:%S'),
            reporte.usuario.username,
            reporte.codigo_arancelario or '--',
            (reporte.descripcion_clasificacion or '--')[:200],
            dict(reporte.TIPO_ACCION_CHOICES).get(reporte.tipo_accion, reporte.tipo_accion),
            dict(reporte._meta.get_field('resultado_operacion').choices).get(reporte.resultado_operacion, reporte.resultado_operacion),
            reporte.detalles_adicionales or ''
        ]
        data.append(row)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*72, bottomMargin=0.5*72)
    styles = getSampleStyleSheet()
    elements = []
    title = 'Reporte de Trazabilidad Completo' if es_gerente else 'Mi Reporte de Trazabilidad'
    elements.append(Paragraph(title, styles['Heading1']))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    elements.append(table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=reportes_trazabilidad.pdf'
    response.write(pdf)
    return response


# ============= USER PROFILE VIEWS =============

@login_required
def edit_profile(request):
    """View to edit user profile"""
    if request.method == 'POST':
        form = EditProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado correctamente.')
            return redirect('central:edit_profile')
        else:
            messages.error(request, 'Hay errores en el formulario. Por favor verifica.')
    else:
        form = EditProfileForm(instance=request.user)
    
    return render(request, 'central/edit_profile.html', {'form': form})


@login_required
def change_password(request):
    """View to change user password"""
    if request.method == 'POST':
        form = ChangePasswordForm(request.POST)
        if form.is_valid():
            # Verify old password
            if request.user.check_password(form.cleaned_data['old_password']):
                # Set new password
                request.user.set_password(form.cleaned_data['new_password'])
                request.user.save()
                
                # Keep user logged in after password change
                update_session_auth_hash(request, request.user)
                
                messages.success(request, 'Tu contraseña ha sido cambiada correctamente.')
                return redirect('central:home')
            else:
                form.add_error('old_password', 'La contraseña actual es incorrecta.')
    else:
        form = ChangePasswordForm()
    
    return render(request, 'central/change_password.html', {'form': form})
    